import xml.etree.ElementTree as ET
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import config

METADATA_FOLDER_PATH = config.metadata_folder_path
METADATA_FILE_NAME = config.metadata_file_name
WORKBOOK_PATH = config.workbook_path

# 🔹 Extract and Save Metadata
def extract_datasource_metadata():
    tree = ET.parse(WORKBOOK_PATH)
    root = tree.getroot()
    datasources = []
    for datasource in root.findall("datasources/datasource"):
        datasource_name = datasource.get("caption")
        connections = datasource.findall("connection")
        for connection in connections:
            connection_details = connection.find(".//connection")
            connection_type = connection_details.get("class")
            file_path = None
            tables = None
            # The data source is a CSV file
            if connection_type == "textscan":
                file_path = fr"{connection_details.get('directory')}/{connection_details.get('filename')}"
                tables = connection.findall("relation")
                for table in tables:
                    tableName = table.get("name")
                    tableName = tableName[:tableName.index(".")]
                    columns = table.findall("columns/column")
                    columnNames = []
                    for column in columns:
                        columnNames.append(column.get("name"))
                    columnNames = ", ".join(columnNames)
                    datasources.append({
                        "Data_Source": datasource_name,
                        "Connection_Info": file_path,
                        "Data_Table_Name": tableName,
                        "Column_Names": columnNames
                    })
            # The data source is an Excel file
            elif connection_type == "excel-direct":
                file_path = connection_details.get('filename')
                tables = connection.findall("relation")
                if tables[0].get("type") == "collection":
                    # There is more than one sheet in the Excel workbook
                    tables = tables[0].findall("relation")
                for table in tables:
                    tableName = table.get("name")
                    columns = table.findall("columns/column")
                    columnNames = []
                    for column in columns:
                        columnNames.append(column.get("name"))
                    columnNames = ", ".join(columnNames)
                    datasources.append({
                        "Data_Source": datasource_name,
                        "Connection_Info": file_path,
                        "Data_Table_Name": tableName,
                        "Column_Names": columnNames
                    })
            # The data source is a hyper file
            elif connection_type == "hyper":
                file_path = connection_details.get("dbname")
                tables = connection.findall("relation")
                table_to_columns = {}
                for table in tables:
                    tableName = table.get("name")
                    table_to_columns[tableName] = []
                columnInfo = connection.findall("metadata-records/metadata-record")
                for c in columnInfo:
                    c_name = c.find("local-name").text[1:-1]
                    c_parent = c.find("parent-name").text[1:-1]
                    table_to_columns.get(c_parent).append(c_name)

                for tableName in table_to_columns:
                    columnNames = ", ".join(table_to_columns[tableName])
                    datasources.append({
                        "Data_Source": datasource_name,
                        "Connection_Info": file_path,
                        "Data_Table_Name": tableName,
                        "Column_Names": columnNames
                    })
    return datasources

def extract_parameter_metadata():
    workbook_path = r"FILL"
    tree = ET.parse(workbook_path)
    root = tree.getroot()
    parameters = []
    parameter_datasource = root.find("datasources/datasource[@name='Parameters']")
    if parameter_datasource is not None:
        column_info = parameter_datasource.find("column")
        caption = column_info.get("caption")
        data_type = column_info.get("datatype")
        format = column_info.get("default-format")
        name = column_info.get("name")
        domain_type = column_info.get("param-domain-type")
        role = column_info.get("role")
        type = column_info.get("type")
        default_value = column_info.get("value")
        range_info = column_info.find("range")
        range_granularity = range_info.get("granularity")
        range_max = range_info.get("max")
        range_min = range_info.get("min")
        parameters.append({
            "Caption": caption,
            "Data Type": data_type,
            "Format": format,
            "Name": name,
            "Domain Type": domain_type,
            "Role": role,
            "Type": type,
            "Default Value": default_value,
            "Range Granularity": range_granularity,
            "Range Min": range_min,
            "Range Max": range_max
        })
    return parameters

def get_mapping():
    column_to_table_mapping = {}
    workbook_path = r"FILL"
    tree = ET.parse(workbook_path)
    root = tree.getroot()
    datasources = root.findall("./datasources/datasource")
    for datasource in datasources:
        for map in datasource.findall("connection/cols/map"):
            column = map.get("key")
            table = map.get("value").split(".")[0][1:-1]
            column_to_table_mapping[column] = table
    return column_to_table_mapping

def extract_visualization_metadata():
    workbook_path = r"FILL"
    tree = ET.parse(workbook_path)
    root = tree.getroot()
    visualizations = []
    for worksheet in root.findall(".//worksheets/worksheet"):
        worksheet_name = worksheet.get("name")
        # the default
        viz_title = "No title"
        viz_title_element = worksheet.find(".//title")
        if viz_title_element is not None:
            viz_title = ""
            for run in viz_title_element.findall(".//run"):
                viz_title = viz_title + run.text

        mark_types = []
        for pane in worksheet.findall(".//panes/pane"):
            mark_types.append(pane.find("mark").get("class"))
        mark_types = ", ".join(mark_types)

        column_content = worksheet.find(".//cols").text
        row_content = worksheet.find(".//rows").text
        reg_ex = r"\]\.\[(.*?)\]"
        columns = "None" if column_content is None else re.findall(reg_ex, column_content)
        rows = "None" if row_content is None else re.findall(reg_ex, row_content)

        visualizations.append({
            "Worksheet Name": worksheet_name,
            "Visualization Title": viz_title,
            "Mark Types": mark_types,
            "Column Shelf": columns,
            "Row Shelf": rows,
            "Filters": [],
            "Measure Values": []
        })
    return visualizations

def adjust_column_widths(dataframe, ws):
    for idx, col in enumerate(dataframe):
        series = dataframe[col]
        max_len = max((
        series.astype(str).map(len).max(),  # len of largest item
        len(str(series.name))  # len of column name/header
        )) + 1  # adding a little extra space
        col_letter = get_column_letter(idx + 1)
        ws.column_dimensions[col_letter].width = max_len

def save_to_excel(datasource_info, sheet_name):
    metadata_file_path = f"{METADATA_FOLDER_PATH}/{METADATA_FILE_NAME}"
    df_datasources = pd.DataFrame(datasource_info)
    df_datasources.to_excel(excel_writer=metadata_file_path, sheet_name=sheet_name, index=False, engine="openpyxl")
    wb = load_workbook(metadata_file_path)
    ws = wb.active
    adjust_column_widths(df_datasources, ws)
    wb.save(metadata_file_path)

def save_to_csv(datasource_info):
    metadata_file_path = f"{METADATA_FOLDER_PATH}/{METADATA_FILE_NAME}"
    df_datasources = pd.DataFrame(datasource_info)
    df_datasources.to_csv(path_or_buf=metadata_file_path, index=False)